"""
Generador de Excel con precios históricos y gráficos de evolución.
Se ejecuta automáticamente tras cada scraping y guarda el .xlsx en el repo.
"""

import csv
from pathlib import Path
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Instala openpyxl: pip install openpyxl")

DATA_DIR  = Path(__file__).parent.parent / "data"
PRICES_CSV = DATA_DIR / "prices.csv"
EXCEL_OUT  = DATA_DIR / "precios_sikaflex.xlsx"

PRODUCTS = ["522", "554", "621"]

# ── Paleta de colores ─────────────────────────────────────────
RED_DARK   = "C0281C"
RED_LIGHT  = "E63B2E"
GREY_DARK  = "2C3E50"
GREY_MID   = "7F8C8D"
GREY_LIGHT = "ECF0F1"
WHITE      = "FFFFFF"
GREEN      = "27AE60"
ORANGE     = "E67E22"

PRODUCT_COLORS = {
    "522": "2980B9",   # azul
    "554": "8E44AD",   # morado
    "621": "27AE60",   # verde
}

def _border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _load_data():
    """Carga el CSV y devuelve lista de dicts ordenada por fecha."""
    rows = []
    if not PRICES_CSV.exists():
        return rows
    with open(PRICES_CSV, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            try:
                r["price"] = float(r["price"])
                rows.append(r)
            except (ValueError, KeyError):
                continue
    return sorted(rows, key=lambda x: x["date"])

def _all_dates(rows):
    return sorted(set(r["date"] for r in rows))

def _all_stores(rows, product=None):
    return sorted(set(
        r["store"] for r in rows
        if product is None or r["product"] == product
    ))

# ══════════════════════════════════════════════════════════════
# HOJA 1 — Todos los precios (tabla completa)
# ══════════════════════════════════════════════════════════════
def _sheet_all_prices(wb, rows):
    ws = wb.create_sheet("📋 Todos los precios")
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "HISTÓRICO COMPLETO DE PRECIOS — SIKAFLEX"
    c.font = Font(bold=True, size=14, color=WHITE)
    c.fill = _header_fill(RED_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Cabeceras
    headers = ["Fecha", "Producto", "Tienda", "Precio (€)", "URL"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = _header_fill(GREY_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    ws.row_dimensions[2].height = 22

    # Datos
    fill_alt = PatternFill("solid", fgColor="F5F6FA")
    for i, r in enumerate(rows, 3):
        row_fill = fill_alt if i % 2 == 0 else None
        vals = [r["date"], r["product"], r["store"], r["price"], r.get("url","")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = _border()
            c.alignment = Alignment(vertical="center")
            if row_fill:
                c.fill = row_fill
            if col == 4:  # precio
                c.number_format = '#,##0.00 "€"'
                c.alignment = Alignment(horizontal="right", vertical="center")

    # Anchos
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════
# HOJA 2 — Media diaria por producto (pivot)
# ══════════════════════════════════════════════════════════════
def _sheet_daily_avg(wb, rows):
    ws = wb.create_sheet("📊 Media diaria")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "MEDIA DE PRECIO DIARIA POR PRODUCTO"
    c.font = Font(bold=True, size=13, color=WHITE)
    c.fill = _header_fill(RED_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Calcular medias
    from statistics import mean
    date_product_prices = defaultdict(lambda: defaultdict(list))
    for r in rows:
        date_product_prices[r["date"]][r["product"]].append(r["price"])

    dates = _all_dates(rows)

    # Cabeceras
    headers = ["Fecha"] + [f"Sikaflex {p}" for p in PRODUCTS]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = _header_fill(PRODUCT_COLORS.get(headers[col-1].replace("Sikaflex ",""), GREY_DARK)
                              if col > 1 else _header_fill(GREY_DARK).fgColor.rgb and GREY_DARK)
        c.fill = _header_fill(GREY_DARK) if col == 1 else _header_fill(PRODUCT_COLORS[PRODUCTS[col-2]])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    ws.row_dimensions[2].height = 22

    fill_alt = PatternFill("solid", fgColor="F5F6FA")
    for i, date in enumerate(dates, 3):
        row_fill = fill_alt if i % 2 == 0 else None
        ws.cell(row=i, column=1, value=date).border = _border()
        if row_fill:
            ws.cell(row=i, column=1).fill = row_fill
        for col, prod in enumerate(PRODUCTS, 2):
            prices = date_product_prices[date][prod]
            c = ws.cell(row=i, column=col,
                        value=round(mean(prices), 2) if prices else None)
            c.number_format = '#,##0.00 "€"'
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _border()
            if row_fill:
                c.fill = row_fill

    ws.column_dimensions["A"].width = 14
    for col in range(2, len(PRODUCTS)+2):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A3"

    # ── Gráfico de evolución ──────────────────────────────────
    if len(dates) >= 2:
        chart = LineChart()
        chart.title = "Evolución del precio medio diario"
        chart.style = 10
        chart.y_axis.title = "Precio (€)"
        chart.x_axis.title = "Fecha"
        chart.height = 14
        chart.width  = 28

        n = len(dates)
        for col, prod in enumerate(PRODUCTS, 2):
            data_ref = Reference(ws, min_col=col, max_col=col,
                                 min_row=2, max_row=2+n)
            chart.add_data(data_ref, titles_from_data=True)

        cats = Reference(ws, min_col=1, min_row=3, max_row=2+n)
        chart.set_categories(cats)

        for i, prod in enumerate(PRODUCTS):
            s = chart.series[i]
            s.graphicalProperties.line.solidFill = PRODUCT_COLORS[prod]
            s.graphicalProperties.line.width = 20000
            s.marker.symbol = "circle"
            s.marker.size   = 5

        ws.add_chart(chart, f"A{len(dates)+5}")

# ══════════════════════════════════════════════════════════════
# HOJAS 3-5 — Una por producto con tabla pivot tienda × fecha
# ══════════════════════════════════════════════════════════════
def _sheet_product(wb, rows, product):
    emoji = {"522": "🔵", "554": "🟣", "621": "🟢"}.get(product, "")
    ws = wb.create_sheet(f"{emoji} Sikaflex {product}")
    ws.sheet_view.showGridLines = False

    prod_rows = [r for r in rows if r["product"] == product]
    dates  = _all_dates(prod_rows)
    stores = _all_stores(prod_rows, product)

    if not dates or not stores:
        ws["A1"] = f"Sin datos para Sikaflex {product}"
        return

    color = PRODUCT_COLORS.get(product, GREY_DARK)

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(dates)+1)}1")
    c = ws["A1"]
    c.value = f"SIKAFLEX {product} — Precio por tienda y fecha"
    c.font = Font(bold=True, size=13, color=WHITE)
    c.fill = _header_fill(color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Cabecera fila 2: "Tienda" + fechas
    ws.cell(row=2, column=1, value="Tienda").font = Font(bold=True, color=WHITE, size=10)
    ws.cell(row=2, column=1).fill = _header_fill(GREY_DARK)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1).border = _border()
    ws.row_dimensions[2].height = 22

    for col, d in enumerate(dates, 2):
        c = ws.cell(row=2, column=col, value=d)
        c.font = Font(bold=True, color=WHITE, size=9)
        c.fill = _header_fill(color)
        c.alignment = Alignment(horizontal="center", rotation=45)
        c.border = _border()

    # Construir lookup {(date, store): price}
    lookup = {}
    for r in prod_rows:
        lookup[(r["date"], r["store"])] = r["price"]

    fill_alt = PatternFill("solid", fgColor="F5F6FA")
    for row_i, store in enumerate(stores, 3):
        row_fill = fill_alt if row_i % 2 == 0 else None
        c = ws.cell(row=row_i, column=1, value=store)
        c.font = Font(bold=True, size=9)
        c.border = _border()
        c.alignment = Alignment(vertical="center")
        if row_fill: c.fill = row_fill

        for col, d in enumerate(dates, 2):
            price = lookup.get((d, store))
            c = ws.cell(row=row_i, column=col, value=price)
            c.number_format = '#,##0.00 "€"'
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _border()
            if row_fill: c.fill = row_fill

    # Anchos
    ws.column_dimensions["A"].width = 26
    for col in range(2, len(dates)+2):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = "B3"

    # ── Gráfico de evolución por tienda ──────────────────────
    if len(dates) >= 2:
        chart = LineChart()
        chart.title = f"Sikaflex {product} — Evolución de precio por tienda"
        chart.style = 10
        chart.y_axis.title = "Precio (€)"
        chart.x_axis.title = "Fecha"
        chart.height = 14
        chart.width  = 30

        for row_i, store in enumerate(stores, 3):
            data_ref = Reference(ws, min_col=2, max_col=1+len(dates),
                                 min_row=row_i, max_row=row_i)
            series = chart.series.__class__()
            chart.add_data(data_ref)
            chart.series[-1].title = SeriesLabel(v=store)

        cats = Reference(ws, min_col=2, max_col=1+len(dates), min_row=2)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{len(stores)+5}")

# ══════════════════════════════════════════════════════════════
# HOJA RESUMEN — Dashboard con precios actuales
# ══════════════════════════════════════════════════════════════
def _sheet_summary(wb, rows):
    ws = wb.create_sheet("🏠 Resumen", 0)
    ws.sheet_view.showGridLines = False

    # Título principal
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "📊 MONITOR DE PRECIOS SIKAFLEX"
    c.font = Font(bold=True, size=16, color=WHITE)
    c.fill = _header_fill(RED_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Fecha actualización
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = f"Última actualización: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(italic=True, size=10, color=GREY_MID)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Últimos precios por producto
    row = 4
    last_date = max(r["date"] for r in rows) if rows else None

    for product in PRODUCTS:
        color = PRODUCT_COLORS[product]
        prod_rows = [r for r in rows if r["product"] == product
                     and r["date"] == last_date]

        # Cabecera producto
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1,
                    value=f"  SIKAFLEX {product} — Precios del {last_date or 'N/A'}")
        c.font = Font(bold=True, size=11, color=WHITE)
        c.fill = _header_fill(color)
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1

        if prod_rows:
            # Sub-cabecera
            for col, h in enumerate(["Tienda", "Precio", "", "", ""], 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = Font(bold=True, size=9, color=WHITE)
                c.fill = _header_fill(GREY_DARK)
                c.border = _border()
            ws.row_dimensions[row].height = 18
            row += 1

            prod_rows_sorted = sorted(prod_rows, key=lambda x: x["price"])
            min_price = prod_rows_sorted[0]["price"]
            max_price = prod_rows_sorted[-1]["price"]

            for r in prod_rows_sorted:
                c1 = ws.cell(row=row, column=1, value=r["store"])
                c1.border = _border()
                c1.font = Font(size=9)

                c2 = ws.cell(row=row, column=2, value=r["price"])
                c2.number_format = '#,##0.00 "€"'
                c2.alignment = Alignment(horizontal="right")
                c2.border = _border()
                c2.font = Font(size=9, bold=True)

                # Marcar min/max
                if r["price"] == min_price:
                    c2.font = Font(bold=True, color=GREEN, size=9)
                    ws.cell(row=row, column=3, value="✅ Más barato").font = Font(color=GREEN, size=8)
                elif r["price"] == max_price:
                    c2.font = Font(bold=True, color=RED_LIGHT, size=9)
                    ws.cell(row=row, column=3, value="🔴 Más caro").font = Font(color=RED_LIGHT, size=8)

                ws.row_dimensions[row].height = 17
                row += 1
        else:
            ws.cell(row=row, column=1, value="Sin datos para esta fecha")
            row += 1

        row += 1  # espacio entre productos

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
def generate_excel():
    rows = _load_data()
    if not rows:
        print("⚠️  No hay datos en prices.csv — Excel no generado.")
        return

    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    _sheet_summary(wb, rows)
    _sheet_daily_avg(wb, rows)
    for product in PRODUCTS:
        _sheet_product(wb, rows, product)
    _sheet_all_prices(wb, rows)

    DATA_DIR.mkdir(exist_ok=True)
    wb.save(EXCEL_OUT)
    print(f"✅ Excel generado: {EXCEL_OUT.name}  ({len(rows)} registros)")
    return EXCEL_OUT

if __name__ == "__main__":
    generate_excel()
